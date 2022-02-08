
import openpyxl
path="C:\\Users\\HP\Desktop\\attendance.xlsx"
att=openpyxl.load_workbook(path)
sheets=att.sheetnames
print(att.active.title)
sh=att[sheets[0]]
print(sh.iter_rows)

